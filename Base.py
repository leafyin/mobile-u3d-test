# encoding=utf-8
import abc
import platform

import openpyxl
from airtest.core.api import *
from poco.drivers.android.uiautomation import AndroidUiautomationPoco


def gen_py():
    """
    根据excel表格生成python对象，写脚本的时候可以运行该方法，把表格内容转为python类对象
    :return:
    """
    prefix = ""
    filepath = "../ResourceConfig.xlsx"
    workbook = openpyxl.load_workbook(filepath)
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row + 1
        print(sheet_name)
        with open(file=f"../gen/{sheet_name}.py", mode="w+", encoding="UTF-8-sig") as f:
            f.write(f"class {sheet_name}:\r\n")
            for i in range(1, max_row):
                if sheet_name == "Sheet":
                    continue
                if sheet_name == "Permission":
                    f.write(f"    {sheet[f'A{i}'].value} = \""
                            f"{prefix}{sheet[f'B{i}'].value}\"\r\n")
                elif sheet_name == "Component":
                    f.write(f"    {sheet[f'A{i}'].value} = \"{sheet[f'B{i}'].value}\"\r\n")
                else:
                    f.write(f"    {sheet[f'A{i}'].value} = \""
                            f"{prefix}{sheet[f'B{i}'].value}\"\r\n")


class Base(metaclass=abc.ABCMeta):

    def __init__(self, deviceid, log_dir=None):
        ST.LOG_FILE = f"{deviceid}.txt"
        ST.OPDELAY = 1
        if log_dir is None:
            system = platform.system()
            if system == 'Windows':
                log_dir = ".\\report\\"
            elif system in ['Darwin', 'Linux']:
                log_dir = "./report/"
        auto_setup(devices=[
            f"Android://127.0.0.1:5037/{deviceid}?cap_method=JAVACAP"
            f"&&ori_method=MINICAPORI"
            f"&&touch_method=MAXTOUCH"
        ], logdir=log_dir)
        self.poco = AndroidUiautomationPoco()
        self.device = self.poco.device
        self.width = self.device.display_info["width"]
        self.height = self.device.display_info["height"]
        # 点亮屏幕
        if not self.device.is_screenon():
            self.poco.device.keyevent(keyname="26")

    @abc.abstractmethod
    def click(self, name):
        pass

    @abc.abstractmethod
    def double_click(self, name):
        pass

    def key_event(self, keycode):
        """
        调用KeyEvent事件
        :param keycode:
        :return:
        """
        self.device.keyevent(str(keycode))

    @abc.abstractmethod
    def set_text(self, name, content):
        pass

    @abc.abstractmethod
    def find_on_vertical(self, obj, length, top_target=None, end_target=None):
        pass

    @abc.abstractmethod
    def find_on_horizontal(self, obj, target, length):
        pass

